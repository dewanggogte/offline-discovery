"""
AC Price Caller - Voice AI Pipeline
====================================
Calls local electronics stores using Sarvam AI (STT/TTS) + LiveKit (telephony)
to enquire about AC prices and collects structured quotes.

Architecture:
  LiveKit Agent (outbound SIP) → Sarvam TTS (Hindi/Hinglish) → Phone Call
  Phone Call audio → Sarvam STT → LLM (extracts price data) → Results DB

Prerequisites:
  1. Sarvam AI API key (https://dashboard.sarvam.ai) - ₹1000 free credits
  2. LiveKit Cloud account (https://cloud.livekit.io) - free tier available
  3. SIP trunk provider (Twilio/Telnyx) with Indian number
  4. OpenAI API key (for LLM-based conversation + extraction)
"""

import asyncio
import json
import os
import sys
import logging
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional

from dotenv import load_dotenv

load_dotenv(".env.local")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler("call_log.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger("ac-price-caller")


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------
@dataclass
class Store:
    name: str
    phone: str  # E.164 format: +91XXXXXXXXXX
    area: str
    city: str


@dataclass
class PriceQuote:
    store: Store
    ac_model: str
    quoted_price: Optional[float] = None
    mrp: Optional[float] = None
    exchange_offer: Optional[str] = None
    installation_included: bool = False
    warranty_info: Optional[str] = None
    availability: Optional[str] = None
    features: Optional[str] = None
    freebies: Optional[str] = None
    additional_notes: Optional[str] = None
    call_duration_sec: float = 0.0
    call_status: str = "pending"  # pending, completed, failed, no_answer, refused
    raw_transcript: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class CallCampaign:
    ac_model: str
    city: str
    stores: list[Store]
    quotes: list[PriceQuote] = field(default_factory=list)
    started_at: str = field(default_factory=lambda: datetime.now().isoformat())
    completed_at: Optional[str] = None


# ---------------------------------------------------------------------------
# Store database (replace with your actual local stores)
# ---------------------------------------------------------------------------
SAMPLE_STORES = [
    Store("Pai International - Jayanagar", "+918042464343", "Jayanagar 2nd Block", "Bangalore"),
    Store("Pai International - Malleshwaram", "+919108444777", "Malleshwaram", "Bangalore"),
    Store("Girias - Koramangala", "+918041571296", "Koramangala", "Bangalore"),
    Store("Girias - Brigade Road", "+919035499066", "Brigade Road", "Bangalore"),
    Store("Viveks - Jayanagar", "+918041461902", "Jayanagar 3rd Block", "Bangalore"),
    Store("Viveks - Indiranagar", "+918041461910", "Indiranagar", "Bangalore"),
    Store("Croma - Koramangala", "+917795838018", "Koramangala", "Bangalore"),
    Store("Croma - Electronic City", "+918867854558", "Electronic City", "Bangalore"),
    Store("Reliance Digital - HSR Layout", "+918071890032", "HSR Layout", "Bangalore"),
    Store("Reliance Digital - Whitefield", "+918692007583", "Whitefield", "Bangalore"),
]

# The AC model we're enquiring about
TARGET_AC = "Samsung 1.5 Ton 5 Star Inverter Split AC (AR18CYNZABE)"


def load_stores(filepath: str = "stores.json") -> list[Store]:
    """Load stores from JSON file, or return sample stores."""
    if os.path.exists(filepath):
        with open(filepath) as f:
            data = json.load(f)
        return [Store(**s) for s in data]
    logger.warning(f"No {filepath} found, using sample stores")
    return SAMPLE_STORES


def save_results(campaign: CallCampaign, filepath: str = "results.json"):
    """Save campaign results to JSON."""
    with open(filepath, "w") as f:
        json.dump(asdict(campaign), f, indent=2, ensure_ascii=False)
    logger.info(f"Results saved to {filepath}")


def print_summary(campaign: CallCampaign):
    """Print a formatted summary of all quotes."""
    print("\n" + "=" * 70)
    print(f"  AC PRICE COMPARISON: {campaign.ac_model}")
    print(f"  City: {campaign.city}")
    print(f"  Stores called: {len(campaign.quotes)}")
    print("=" * 70)

    completed = [q for q in campaign.quotes if q.call_status == "completed" and q.quoted_price]
    failed = [q for q in campaign.quotes if q.call_status != "completed"]

    if completed:
        completed.sort(key=lambda q: q.quoted_price)
        print(f"\n  ✅ Successful quotes: {len(completed)}")
        print(f"  ❌ Failed/No answer: {len(failed)}")
        print(f"\n  {'Store':<30} {'Area':<20} {'Price':>10} {'Exchange':>15}")
        print("  " + "-" * 75)
        for q in completed:
            exchange = q.exchange_offer or "N/A"
            print(f"  {q.store.name:<30} {q.store.area:<20} ₹{q.quoted_price:>8,.0f} {exchange:>15}")

        best = completed[0]
        worst = completed[-1]
        print(f"\n  🏆 BEST PRICE: {best.store.name} — ₹{best.quoted_price:,.0f}")
        if len(completed) > 1:
            savings = worst.quoted_price - best.quoted_price
            print(f"  💰 MAX SAVINGS: ₹{savings:,.0f} vs highest quote")
    else:
        print("\n  No successful quotes obtained.")

    if failed:
        print(f"\n  Failed calls:")
        for q in failed:
            print(f"    - {q.store.name}: {q.call_status}")

    print("\n" + "=" * 70)


def export_to_excel(campaign: CallCampaign, filepath: str = "results.xlsx"):
    """Export campaign results to a formatted Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "AC Price Quotes"

    headers = [
        "Store", "Area", "Price", "MRP", "Exchange Offer",
        "Installation", "Warranty", "Features", "Freebies",
        "Availability", "Call Status", "Notes",
    ]
    ws.append(headers)

    # Bold header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Add data rows
    for q in campaign.quotes:
        ws.append([
            q.store.name,
            q.store.area,
            q.quoted_price,
            q.mrp,
            q.exchange_offer or "",
            "Yes" if q.installation_included else "No",
            q.warranty_info or "",
            q.features or "",
            q.freebies or "",
            q.availability or "",
            q.call_status,
            q.additional_notes or "",
        ])

    # Auto-column-width
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Highlight best price in green
    completed = [q for q in campaign.quotes if q.call_status == "completed" and q.quoted_price]
    if completed:
        best_price = min(q.quoted_price for q in completed)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        price_col = 3  # Column C
        for row_idx in range(2, len(campaign.quotes) + 2):
            cell = ws.cell(row=row_idx, column=price_col)
            if cell.value == best_price:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = green_fill

    wb.save(filepath)
    logger.info(f"Excel report saved to {filepath}")


async def run_campaign(ac_model: str = TARGET_AC, city: str = "Bangalore"):
    """
    Main entry point: orchestrates calling all stores.
    """
    # Validate environment
    required_vars = ["SARVAM_API_KEY", "LIVEKIT_URL", "LIVEKIT_API_KEY", "LIVEKIT_API_SECRET", "LLM_BASE_URL"]
    missing = [v for v in required_vars if not os.getenv(v)]
    if missing:
        logger.error(f"Missing environment variables: {', '.join(missing)}")
        logger.error("Copy .env.example to .env.local and fill in your keys.")
        sys.exit(1)

    sip_trunk_id = os.getenv("SIP_OUTBOUND_TRUNK_ID")
    if not sip_trunk_id:
        logger.error("SIP_OUTBOUND_TRUNK_ID is required for outbound calling")
        logger.error("Set up a Twilio/Telnyx SIP trunk and register it with LiveKit")
        sys.exit(1)

    stores = load_stores()
    campaign = CallCampaign(ac_model=ac_model, city=city, stores=stores)

    logger.info(f"Starting campaign: {ac_model} in {city}")
    logger.info(f"Calling {len(stores)} stores...")

    # Import the caller (separated for cleaner architecture)
    from caller import make_price_enquiry_call

    for i, store in enumerate(stores, 1):
        logger.info(f"\n--- Call {i}/{len(stores)}: {store.name} ({store.phone}) ---")

        quote = await make_price_enquiry_call(
            store=store,
            ac_model=ac_model,
            sip_trunk_id=sip_trunk_id,
        )
        campaign.quotes.append(quote)

        # Save incrementally (in case of crashes)
        save_results(campaign)

        # Delay between calls to be polite and avoid rate limits
        if i < len(stores):
            delay = 10
            logger.info(f"Waiting {delay}s before next call...")
            await asyncio.sleep(delay)

    campaign.completed_at = datetime.now().isoformat()
    save_results(campaign)
    print_summary(campaign)
    export_to_excel(campaign)

    return campaign


if __name__ == "__main__":
    asyncio.run(run_campaign())
